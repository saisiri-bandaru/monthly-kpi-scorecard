#!/usr/bin/env python3
"""Build Northline_Monthly_KPI_Scorecard.xlsx — ops + finance RAG scorecard portfolio sample."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

yellow = PatternFill("solid", fgColor="FFF2CC")
header_fill = PatternFill("solid", fgColor="1F4E79")
section_fill = PatternFill("solid", fgColor="D6E3F0")
green_fill = PatternFill("solid", fgColor="C6EFCE")
amber_fill = PatternFill("solid", fgColor="FFE699")
red_fill = PatternFill("solid", fgColor="F8CBAD")
tile_fill = PatternFill("solid", fgColor="E9EDF4")
light_gray = PatternFill("solid", fgColor="F5F5F5")
input_font = Font(name="Calibri", size=11, color="0000FF")
black = Font(name="Calibri", size=11, color="000000")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
title_font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
section_font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
bold = Font(name="Calibri", size=11, bold=True)
bold_black = Font(name="Calibri", size=11, bold=True, color="000000")
italic_grey = Font(name="Calibri", size=10, italic=True, color="666666")
small_grey = Font(name="Calibri", size=9, italic=True, color="666666")
link_font = Font(name="Calibri", size=11, color="0563C1", underline="single")
rag_font = Font(name="Calibri", size=11, bold=True)
thin = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
money = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'
pct = "0.0%"
days_fmt = "0.0"
fte_fmt = "0.0"
num = "#,##0.0"

MONTHS = [
    "Jan-26", "Feb-26", "Mar-26", "Apr-26", "May-26", "Jun-26",
    "Jul-26", "Aug-26", "Sep-26", "Oct-26", "Nov-26", "Dec-26",
]

# Align with the WC / HC samples where it is natural to do so.
REV_TGT = [71000, 68000, 73500, 75000, 77500, 79000, 80500, 79800, 77200, 82000, 87500, 94000]
REV_ACT = [70200, 67400, 74100, 73800, 78200, 80100, 79800, 80500, 76500, 83400, 89200, 95800]
GM_TGT = [0.422, 0.422, 0.424, 0.425, 0.426, 0.427, 0.427, 0.426, 0.425, 0.428, 0.430, 0.432]
GM_ACT = [0.418, 0.415, 0.422, 0.425, 0.428, 0.430, 0.427, 0.424, 0.421, 0.426, 0.432, 0.435]
EBITDA_TGT = [round(r * 0.145) for r in REV_TGT]
EBITDA_ACT = [round(a * m) for a, m in zip(REV_ACT, [0.138, 0.136, 0.144, 0.147, 0.149, 0.151, 0.146, 0.143, 0.141, 0.148, 0.152, 0.154])]
VOL_TGT = [round(r / 12) for r in REV_TGT]
VOL_ACT = [round(a / 12 * k) for a, k in zip(REV_ACT, [0.99, 0.985, 1.01, 0.98, 1.015, 1.02, 0.99, 1.01, 0.985, 1.02, 1.025, 1.03])]
OTIF_TGT = [0.970] * 12
OTIF_ACT = [0.962, 0.951, 0.974, 0.978, 0.969, 0.975, 0.948, 0.961, 0.970, 0.976, 0.981, 0.964]
SCRAP_TGT = [0.018] * 12
SCRAP_ACT = [0.021, 0.024, 0.019, 0.017, 0.018, 0.016, 0.022, 0.020, 0.019, 0.017, 0.015, 0.016]
DSO_TGT = [42.0] * 12
DSO_ACT = [44.2, 45.1, 43.6, 42.9, 41.4, 42.1, 43.0, 44.2, 42.5, 41.2, 40.0, 41.6]
DIO_ACT = [60.5, 61.2, 59.4, 58.1, 57.0, 56.2, 56.8, 57.5, 58.6, 55.4, 52.1, 53.8]
DPO_ACT = [36.2, 37.0, 38.1, 39.0, 39.8, 38.4, 37.2, 36.5, 37.8, 38.6, 39.5, 40.8]
CCC_TGT = [62.0] * 12
CCC_ACT = [round(d + i - p, 1) for d, i, p in zip(DSO_ACT, DIO_ACT, DPO_ACT)]
HC_TGT = [286.5, 287.8, 289.2, 291.0, 292.4, 294.1, 297.0, 298.4, 299.1, 302.3, 303.5, 304.8]
HC_ACT = [286.5, 285.9, 288.4, 289.6, 291.8, 292.5, 295.6, 296.8, 297.9, 300.7, 302.8, 304.1]
ATTR_TGT = [0.12] * 12
ATTR_ACT = [0.118, 0.152, 0.109, 0.101, 0.097, 0.141, 0.112, 0.134, 0.108, 0.099, 0.095, 0.102]
OCF_TGT = [round(r * 0.08) for r in REV_TGT]
OCF_ACT = [5100, 4200, 6800, 5900, 7400, 7100, 5600, 6200, 4800, 7900, 9100, 8600]
SGA_TGT = [0.185] * 12
SGA_ACT = [0.191, 0.194, 0.186, 0.184, 0.181, 0.179, 0.188, 0.187, 0.190, 0.182, 0.178, 0.176]

# FY25 monthly (for YoY on a subset)
REV_FY25 = [66000, 64000, 69000, 70500, 72000, 73500, 75000, 74200, 72800, 76000, 80500, 87000]
VOL_FY25 = [round(r / 12) for r in REV_FY25]

ASSUMP = "01_Assumptions"
ACT = "02_Actuals"
SC = "03_Scorecard"
TR = "04_Trends"
DASH = "05_Dashboard"

# kpi_id used as row offset. agg: SUM / AVG / LAST
# direction: H higher-better, L lower-better, R range
# nf: number format
KPIS = [
    dict(name="Net revenue", unit="$000s", direction="H", amber=0.03, green=0.02, agg="SUM", nf=money, tgt=REV_TGT, act=REV_ACT),
    dict(name="Gross margin %", unit="%", direction="H", amber=0.015, green=0.008, agg="AVG", nf=pct, tgt=GM_TGT, act=GM_ACT),
    dict(name="EBITDA", unit="$000s", direction="H", amber=0.05, green=0.02, agg="SUM", nf=money, tgt=EBITDA_TGT, act=EBITDA_ACT),
    dict(name="Volume", unit="000 cases", direction="H", amber=0.03, green=0.02, agg="SUM", nf=num, tgt=VOL_TGT, act=VOL_ACT),
    dict(name="OTIF %", unit="%", direction="H", amber=0.02, green=0.01, agg="AVG", nf=pct, tgt=OTIF_TGT, act=OTIF_ACT),
    dict(name="Scrap %", unit="%", direction="L", amber=0.15, green=0.05, agg="AVG", nf=pct, tgt=SCRAP_TGT, act=SCRAP_ACT),
    dict(name="DSO", unit="days", direction="L", amber=0.08, green=0.03, agg="LAST", nf=days_fmt, tgt=DSO_TGT, act=DSO_ACT),
    dict(name="CCC", unit="days", direction="L", amber=0.08, green=0.03, agg="LAST", nf=days_fmt, tgt=CCC_TGT, act=CCC_ACT),
    dict(name="Headcount", unit="FTE", direction="R", amber=0.03, green=0.015, agg="LAST", nf=fte_fmt, tgt=HC_TGT, act=HC_ACT),
    dict(name="Attrition % (ann.)", unit="%", direction="L", amber=0.15, green=0.05, agg="AVG", nf=pct, tgt=ATTR_TGT, act=ATTR_ACT),
    dict(name="Operating cash flow", unit="$000s", direction="H", amber=0.10, green=0.04, agg="SUM", nf=money, tgt=OCF_TGT, act=OCF_ACT),
    dict(name="SG&A % of sales", unit="%", direction="L", amber=0.04, green=0.015, agg="AVG", nf=pct, tgt=SGA_TGT, act=SGA_ACT),
]
K = len(KPIS)  # 12


def style_input(cell):
    cell.fill = yellow
    cell.font = input_font
    cell.border = thin
    cell.alignment = Alignment(horizontal="center")


def style_formula(cell, key=False):
    cell.font = bold_black if key else black
    cell.border = thin
    cell.alignment = Alignment(horizontal="center")
    if key:
        cell.fill = green_fill


def style_header_cell(cell, value):
    cell.value = value
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def month_headers(ws, row, start_col=3, include_fy=True, fy_label="FY26"):
    for i, m in enumerate(MONTHS):
        style_header_cell(ws.cell(row, start_col + i), m)
    if include_fy:
        style_header_cell(ws.cell(row, start_col + 12), fy_label)


def landscape(ws, fit_height=0):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = fit_height
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
    ws.sheet_view.showGridLines = False
    ws.page_setup.horizontalCentered = True
    ws.oddFooter.left.text = "Northline Consumer Products  |  fictional sample"
    ws.oddFooter.right.text = "Page &P of &N"


def col(i):
    return get_column_letter(3 + i)


def shade_section(ws, row, until=16):
    ws.cell(row, 2).fill = section_fill
    ws.cell(row, 2).font = section_font
    for c in range(3, until):
        ws.cell(row, c).fill = section_fill


def rag_formula(actual, target, dir_cell, amber_cell, green_cell):
    """Excel RAG using H / L / R from the KPI register."""
    return (
        f'IF({target}=0,"—",'
        f'IF({dir_cell}="H",'
        f'IF({actual}>={target},"Green",IF({actual}>={target}*(1-{amber_cell}),"Amber","Red")),'
        f'IF({dir_cell}="L",'
        f'IF({actual}<={target},"Green",IF({actual}<={target}*(1+{amber_cell}),"Amber","Red")),'
        f'IF(ABS({actual}/{target}-1)<={green_cell},"Green",'
        f'IF(ABS({actual}/{target}-1)<={amber_cell},"Amber","Red")))))'
    )


wb = Workbook()

# ========== 00_Cover ==========
ws = wb.active
ws.title = "00_Cover"
ws.sheet_properties.tabColor = "1F4E79"
set_col_widths(ws, [4, 92])
landscape(ws, fit_height=1)

ws["B2"] = "Northline Consumer Products"
ws["B2"].font = title_font
ws["B3"] = "Monthly KPI scorecard  ·  ops + finance  ·  actual vs target with RAG"
ws["B3"].font = section_font
ws["B4"] = "FY2026 monthly  ·  12 KPIs  ·  fictional CPG company"
ws["B4"].font = italic_grey

ws["B6"] = "What this file is"
ws["B6"].font = bold
ws["B7"] = (
    "A twelve-month operating and finance scorecard for a fictional mid-size CPG company. "
    "Targets and RAG thresholds live on Assumptions; monthly actuals are inputs; "
    "the scorecard, YTD, trends, and dashboard are formulas (Green / Amber / Red)."
)
ws["B7"].alignment = Alignment(wrap_text=True)
ws.row_dimensions[7].height = 48

ws["B9"] = "How to use"
ws["B9"].font = bold
ws["B10"] = "1. Edit yellow cells on 01_Assumptions (KPI targets, direction H/L/R, amber/green bands, reporting month)."
ws["B11"] = "2. Paste or type monthly actuals on 02_Actuals (yellow)."
ws["B12"] = "3. Read actual vs target, variance, and RAG on 03_Scorecard — nothing on that tab is hardcoded."
ws["B13"] = "4. Read MoM, rolling 3-month, YTD, and vs-target trends on 04_Trends."
ws["B14"] = "5. Use 05_Dashboard as the one-pager (traffic-light tiles + charts)."

ws["B16"] = "File conventions"
ws["B16"].font = bold
ws["B17"] = "Yellow cells with blue font = inputs. Black font = formulas. Green cells = key outputs."
ws["B17"].fill = yellow
ws["B17"].font = Font(name="Calibri", size=11, color="0000FF", bold=True)
ws["B18"] = "RAG: Green = meet/beat (or inside the range). Amber = miss inside the band. Red = miss beyond the band."
ws["B19"] = "All figures are fictional. There is no employer data in this file. Dollar figures in $000s."

ws["B21"] = "Portfolio"
ws["B21"].font = bold
ws["B22"] = "Sai Siri Bandaru — Financial Analyst | FP&A | forecasting, variance analysis, Excel"
ws["B23"] = "https://github.com/saisiri-bandaru"
ws["B23"].font = link_font

# ========== 01_Assumptions ==========
ws = wb.create_sheet(ASSUMP)
ws.sheet_properties.tabColor = "F7C948"
set_col_widths(ws, [4, 28, 14, 12, 14, 14, 14, 18, 28] + [12] * 12)
landscape(ws)
ws.freeze_panes = "C22"

ws["B2"] = "Assumptions"
ws["B2"].font = title_font
ws["B3"] = "Yellow + blue = inputs. Targets, direction, and RAG bands drive the scorecard."
ws["B3"].font = italic_grey

ws["B5"] = "Reporting month (1 = Jan-26 … 12 = Dec-26)"
ws["C5"] = 9
style_input(ws["C5"]); ws["C5"].number_format = "0"
ws["D5"] = '=INDEX({"Jan-26","Feb-26","Mar-26","Apr-26","May-26","Jun-26","Jul-26","Aug-26","Sep-26","Oct-26","Nov-26","Dec-26"},1,C5)'
style_formula(ws["D5"])
ws["E5"] = "Scorecard as-of and dashboard INDEX to this month. Sample is parked in Sep so YTD is visible."
ws["E5"].font = small_grey

ws["B7"] = "RAG rules"
ws["B7"].font = section_font
ws["B7"].fill = section_fill
ws["B8"] = "H  (higher-better): Green if actual ≥ target; Amber if actual ≥ target × (1 − amber band); else Red."
ws["B9"] = "L  (lower-better): Green if actual ≤ target; Amber if actual ≤ target × (1 + amber band); else Red."
ws["B10"] = "R  (range, e.g. headcount): Green if |actual/target − 1| ≤ green band; Amber if ≤ amber band; else Red."
for r in range(8, 11):
    ws.cell(r, 2).font = small_grey

ws["B12"] = "KPI register"
shade_section(ws, 12, until=10)
reg_headers = ["KPI", "Unit", "Direction (H/L/R)", "Amber band", "Green band (R)", "Aggregation", "Polarity"]
for i, h in enumerate(reg_headers):
    style_header_cell(ws.cell(13, 2 + i), h)

polarity = {
    "H": "Higher is better",
    "L": "Lower is better",
    "R": "Hit the plan (range)",
}
for i, k in enumerate(KPIS):
    r = 14 + i
    ws.cell(r, 2, k["name"]).font = bold
    ws.cell(r, 2).border = thin
    ws.cell(r, 3, k["unit"]).border = thin
    c = ws.cell(r, 4, k["direction"]); style_input(c)
    c = ws.cell(r, 5, k["amber"]); style_input(c); c.number_format = pct
    c = ws.cell(r, 6, k["green"]); style_input(c); c.number_format = pct
    ws.cell(r, 7, k["agg"]).border = thin
    ws.cell(r, 8, polarity[k["direction"]]).border = thin

ws["B27"] = "Monthly targets — FY2026 (inputs)"
shade_section(ws, 27)
month_headers(ws, 28)
# KPI target rows 29-40
for i, k in enumerate(KPIS):
    r = 29 + i
    ws.cell(r, 2, k["name"]).border = thin
    for m, v in enumerate(k["tgt"]):
        cell = ws.cell(r, 3 + m, v)
        style_input(cell)
        cell.number_format = k["nf"]
    # FY aggregation
    if k["agg"] == "SUM":
        f = f"=SUM(C{r}:N{r})"
    elif k["agg"] == "AVG":
        f = f"=AVERAGE(C{r}:N{r})"
    else:
        f = f"=N{r}"
    cell = ws.cell(r, 15, f)
    style_formula(cell, key=True)
    cell.number_format = k["nf"]

ws["B42"] = "FY25 actuals (for YoY on Trends) — revenue and volume only"
shade_section(ws, 42)
month_headers(ws, 43, fy_label="FY25")
ws["B44"] = "Net revenue FY25"
for m, v in enumerate(REV_FY25):
    cell = ws.cell(44, 3 + m, v); style_input(cell); cell.number_format = money
ws["O44"] = "=SUM(C44:N44)"; style_formula(ws["O44"], key=True); ws["O44"].number_format = money
ws["B45"] = "Volume FY25"
for m, v in enumerate(VOL_FY25):
    cell = ws.cell(45, 3 + m, v); style_input(cell); cell.number_format = num
ws["O45"] = "=SUM(C45:N45)"; style_formula(ws["O45"]); ws["O45"].number_format = num

ws["B47"] = "Aggregation: SUM for flow KPIs (revenue, EBITDA, volume, cash). AVG for rates. LAST for stocks (DSO, CCC, headcount)."
ws["B47"].font = small_grey

# ========== 02_Actuals ==========
ws = wb.create_sheet(ACT)
ws.sheet_properties.tabColor = "5B9BD5"
set_col_widths(ws, [4, 28] + [12] * 13)
landscape(ws)
ws.freeze_panes = "C6"

ws["B2"] = "Monthly actuals — FY2026"
ws["B2"].font = title_font
ws["B3"] = "Yellow + blue = inputs. Paste a new month here; scorecard, trends, and dashboard recalculate."
ws["B3"].font = italic_grey

ws.cell(5, 2, "Actuals")
shade_section(ws, 5)
month_headers(ws, 6)
for i, k in enumerate(KPIS):
    r = 7 + i
    ws.cell(r, 2, k["name"]).border = thin
    ws.cell(r, 2).font = bold
    for m, v in enumerate(k["act"]):
        cell = ws.cell(r, 3 + m, v)
        style_input(cell)
        cell.number_format = k["nf"]
    if k["agg"] == "SUM":
        f = f"=SUM(C{r}:N{r})"
    elif k["agg"] == "AVG":
        f = f"=AVERAGE(C{r}:N{r})"
    else:
        f = f"=N{r}"
    cell = ws.cell(r, 15, f)
    style_formula(cell, key=True)
    cell.number_format = k["nf"]

ws["B20"] = "YTD (through reporting month) — formulas"
shade_section(ws, 20)
style_header_cell(ws.cell(21, 2), "KPI")
style_header_cell(ws.cell(21, 3), "YTD actual")
style_header_cell(ws.cell(21, 4), "YTD target")
style_header_cell(ws.cell(21, 5), "YTD var")
style_header_cell(ws.cell(21, 6), "Aggregation")

for i, k in enumerate(KPIS):
    r = 22 + i
    ar = 7 + i
    tr = 29 + i
    ws.cell(r, 2, k["name"]).border = thin
    n = f"'{ASSUMP}'!$C$5"
    if k["agg"] == "SUM":
        fa = f"=SUM(OFFSET(C{ar},0,0,1,{n}))"
        ft = f"=SUM(OFFSET('{ASSUMP}'!C{tr},0,0,1,{n}))"
    elif k["agg"] == "AVG":
        fa = f"=AVERAGE(OFFSET(C{ar},0,0,1,{n}))"
        ft = f"=AVERAGE(OFFSET('{ASSUMP}'!C{tr},0,0,1,{n}))"
    else:
        fa = f"=INDEX(C{ar}:N{ar},1,{n})"
        ft = f"=INDEX('{ASSUMP}'!C{tr}:N{tr},1,{n})"
    cell = ws.cell(r, 3, fa); style_formula(cell, key=True); cell.number_format = k["nf"]
    cell = ws.cell(r, 4, ft); style_formula(cell); cell.number_format = k["nf"]
    cell = ws.cell(r, 5, f"=C{r}-D{r}"); style_formula(cell); cell.number_format = k["nf"]
    ws.cell(r, 6, k["agg"]).border = thin

ws["B35"] = "OFFSET/INDEX YTD respects the reporting month on Assumptions. LAST KPIs use the as-of month, not a sum."
ws["B35"].font = small_grey

# ========== 03_Scorecard ==========
ws = wb.create_sheet(SC)
ws.sheet_properties.tabColor = "70AD47"
set_col_widths(ws, [4, 24, 14, 14, 14, 12, 12, 14, 14, 12, 14] + [11] * 12)
landscape(ws)
ws.freeze_panes = "C8"

ws["B2"] = "Scorecard — actual vs target and RAG"
ws["B2"].font = title_font
ws["B3"] = "Every value on this sheet is a formula. RAG uses the H / L / R rules and bands on Assumptions."
ws["B3"].font = italic_grey
ws["B4"] = "As-of month"
ws["C4"] = f"='{ASSUMP}'!D5"
style_formula(ws["C4"], key=True)

# As-of month detail
ws.cell(6, 2, "As-of month scorecard")
shade_section(ws, 6, until=12)
headers = ["KPI", "Actual", "Target", "Var", "Var %", "RAG", "YTD actual", "YTD target", "YTD var", "YTD RAG"]
for i, h in enumerate(headers):
    style_header_cell(ws.cell(7, 2 + i), h)

for i, k in enumerate(KPIS):
    r = 8 + i
    ar = 7 + i          # actuals row
    tr = 29 + i         # target row on assumptions
    ytd_r = 22 + i      # YTD row on actuals
    dir_c = f"'{ASSUMP}'!$D${14+i}"
    amb_c = f"'{ASSUMP}'!$E${14+i}"
    grn_c = f"'{ASSUMP}'!$F${14+i}"
    n = f"'{ASSUMP}'!$C$5"

    ws.cell(r, 2, f"='{ASSUMP}'!B{14+i}").font = bold
    ws.cell(r, 2).border = thin

    cell = ws.cell(r, 3, f"=INDEX('{ACT}'!C{ar}:N{ar},1,{n})")
    style_formula(cell, key=True); cell.number_format = k["nf"]
    cell = ws.cell(r, 4, f"=INDEX('{ASSUMP}'!C{tr}:N{tr},1,{n})")
    style_formula(cell); cell.number_format = k["nf"]
    cell = ws.cell(r, 5, f"=C{r}-D{r}")
    style_formula(cell); cell.number_format = k["nf"]
    cell = ws.cell(r, 6, f"=IF(D{r}=0,\"—\",C{r}/D{r}-1)")
    style_formula(cell); cell.number_format = pct
    cell = ws.cell(r, 7, "=" + rag_formula(f"C{r}", f"D{r}", dir_c, amb_c, grn_c))
    style_formula(cell); cell.font = rag_font

    cell = ws.cell(r, 8, f"='{ACT}'!C{ytd_r}")
    style_formula(cell); cell.number_format = k["nf"]
    cell = ws.cell(r, 9, f"='{ACT}'!D{ytd_r}")
    style_formula(cell); cell.number_format = k["nf"]
    cell = ws.cell(r, 10, f"=H{r}-I{r}")
    style_formula(cell); cell.number_format = k["nf"]
    cell = ws.cell(r, 11, "=" + rag_formula(f"H{r}", f"I{r}", dir_c, amb_c, grn_c))
    style_formula(cell); cell.font = rag_font

# RAG colors on as-of + YTD
for col_letter in ("G", "K"):
    rng = f"{col_letter}8:{col_letter}19"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Green"'], fill=green_fill))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Amber"'], fill=amber_fill))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Red"'], fill=red_fill))

# Monthly RAG heatmap
ws.cell(21, 2, "Monthly RAG heatmap")
shade_section(ws, 21)
month_headers(ws, 22, include_fy=False)
for i, k in enumerate(KPIS):
    r = 23 + i
    ar = 7 + i
    tr = 29 + i
    dir_c = f"'{ASSUMP}'!$D${14+i}"
    amb_c = f"'{ASSUMP}'!$E${14+i}"
    grn_c = f"'{ASSUMP}'!$F${14+i}"
    ws.cell(r, 2, k["name"]).border = thin
    ws.cell(r, 2).font = bold
    for m in range(12):
        cl = col(m)
        actual = f"'{ACT}'!{cl}{ar}"
        target = f"'{ASSUMP}'!{cl}{tr}"
        cell = ws.cell(r, 3 + m, "=" + rag_formula(actual, target, dir_c, amb_c, grn_c))
        style_formula(cell)
        cell.font = rag_font

ws.conditional_formatting.add(
    "C23:N34",
    CellIsRule(operator="equal", formula=['"Green"'], fill=green_fill),
)
ws.conditional_formatting.add(
    "C23:N34",
    CellIsRule(operator="equal", formula=['"Amber"'], fill=amber_fill),
)
ws.conditional_formatting.add(
    "C23:N34",
    CellIsRule(operator="equal", formula=['"Red"'], fill=red_fill),
)

# Monthly variance % grid (formulas)
ws.cell(36, 2, "Monthly variance %  (actual / target − 1)")
shade_section(ws, 36)
month_headers(ws, 37, include_fy=False)
for i, k in enumerate(KPIS):
    r = 38 + i
    ar = 7 + i
    tr = 29 + i
    ws.cell(r, 2, k["name"]).border = thin
    for m in range(12):
        cl = col(m)
        cell = ws.cell(r, 3 + m, f"=IF('{ASSUMP}'!{cl}{tr}=0,\"—\",'{ACT}'!{cl}{ar}/'{ASSUMP}'!{cl}{tr}-1)")
        style_formula(cell)
        cell.number_format = pct

ws["B51"] = "Green / Amber / Red is computed, not typed. Change a target or an actual and the heatmap moves."
ws["B51"].font = small_grey

# ========== 04_Trends ==========
ws = wb.create_sheet(TR)
ws.sheet_properties.tabColor = "ED7D31"
set_col_widths(ws, [4, 28] + [12] * 13)
landscape(ws)
ws.freeze_panes = "C7"

ws["B2"] = "Trends — MoM, rolling 3-month, YTD, vs target"
ws["B2"].font = title_font
ws["B3"] = "All formulas off 02_Actuals and 01_Assumptions. In-cell bars are a simple sparkline substitute."
ws["B3"].font = italic_grey

# Block 1: actuals echo (for charts)
ws.cell(5, 2, "Actuals (echo)")
shade_section(ws, 5)
month_headers(ws, 6, include_fy=False)
for i, k in enumerate(KPIS):
    r = 7 + i
    ws.cell(r, 2, k["name"]).border = thin
    for m in range(12):
        cl = col(m)
        cell = ws.cell(r, 3 + m, f"='{ACT}'!{cl}{7+i}")
        style_formula(cell)
        cell.number_format = k["nf"]

ws.cell(20, 2, "Month-on-month change")
shade_section(ws, 20)
month_headers(ws, 21, include_fy=False)
for i, k in enumerate(KPIS):
    r = 22 + i
    src = 7 + i
    ws.cell(r, 2, k["name"]).border = thin
    ws.cell(r, 3, "—").alignment = Alignment(horizontal="center")
    ws.cell(r, 3).border = thin
    for m in range(1, 12):
        cl = col(m)
        prev = col(m - 1)
        if k["nf"] in (pct,):
            f = f"={cl}{src}-{prev}{src}"
            nf = pct
        else:
            f = f"=IF({prev}{src}=0,\"—\",{cl}{src}/{prev}{src}-1)"
            nf = pct
        cell = ws.cell(r, 3 + m, f)
        style_formula(cell)
        cell.number_format = nf

ws.cell(35, 2, "Rolling 3-month average")
shade_section(ws, 35)
month_headers(ws, 36, include_fy=False)
for i, k in enumerate(KPIS):
    r = 37 + i
    src = 7 + i
    ws.cell(r, 2, k["name"]).border = thin
    for m in range(12):
        cl = col(m)
        if m == 0:
            f = f"={cl}{src}"
        elif m == 1:
            f = f"=AVERAGE(C{src}:{cl}{src})"
        else:
            a = col(m - 2)
            f = f"=AVERAGE({a}{src}:{cl}{src})"
        cell = ws.cell(r, 3 + m, f)
        style_formula(cell)
        cell.number_format = k["nf"]

ws.cell(50, 2, "Variance vs monthly target")
shade_section(ws, 50)
month_headers(ws, 51, include_fy=False)
for i, k in enumerate(KPIS):
    r = 52 + i
    ws.cell(r, 2, k["name"]).border = thin
    for m in range(12):
        cl = col(m)
        cell = ws.cell(r, 3 + m, f"='{ACT}'!{cl}{7+i}-'{ASSUMP}'!{cl}{29+i}")
        style_formula(cell)
        cell.number_format = k["nf"]

ws.cell(65, 2, "YoY revenue and volume vs FY25")
shade_section(ws, 65)
month_headers(ws, 66, include_fy=False)
ws["B67"] = "Revenue YoY %"
ws["B68"] = "Volume YoY %"
for m in range(12):
    cl = col(m)
    cell = ws.cell(67, 3 + m, f"=IF('{ASSUMP}'!{cl}44=0,\"—\",'{ACT}'!{cl}7/'{ASSUMP}'!{cl}44-1)")
    style_formula(cell); cell.number_format = pct
    cell = ws.cell(68, 3 + m, f"=IF('{ASSUMP}'!{cl}45=0,\"—\",'{ACT}'!{cl}10/'{ASSUMP}'!{cl}45-1)")
    style_formula(cell); cell.number_format = pct

ws.cell(70, 2, "In-cell trend bars (revenue actual)")
shade_section(ws, 70)
month_headers(ws, 71, include_fy=False)
ws["B72"] = "Revenue bar"
for m in range(12):
    cl = col(m)
    cell = ws.cell(72, 3 + m, f'=REPT("█",MAX(0,ROUND({cl}7/8000,0)))')
    style_formula(cell)
    cell.font = Font(name="Calibri", size=8, color="1F4E79")
ws["B73"] = "OTIF bar"
for m in range(12):
    cl = col(m)
    cell = ws.cell(73, 3 + m, f'=REPT("█",MAX(0,ROUND({cl}11*40,0)))')
    style_formula(cell)
    cell.font = Font(name="Calibri", size=8, color="1F4E79")
ws["B74"] = "CCC bar"
for m in range(12):
    cl = col(m)
    cell = ws.cell(74, 3 + m, f'=REPT("█",MAX(0,ROUND({cl}14/8,0)))')
    style_formula(cell)
    cell.font = Font(name="Calibri", size=8, color="1F4E79")

ws["B76"] = "Revenue bar is scaled by $8m; OTIF by 2.5 pts per block; CCC by 8 days per block. Cosmetic only."
ws["B76"].font = small_grey

# ========== 05_Dashboard ==========
ws = wb.create_sheet(DASH)
ws.sheet_properties.tabColor = "1F4E79"
set_col_widths(ws, [4, 20, 14, 12, 3, 20, 14, 12, 3, 20, 14, 12, 14, 14])
landscape(ws, fit_height=1)

ws["B2"] = "KPI dashboard"
ws["B2"].font = title_font
ws["B3"] = "As-of month"
ws["C3"] = f"='{ASSUMP}'!D5"
style_formula(ws["C3"], key=True)
ws["E3"] = "Tiles pull the as-of month from the scorecard. RAG is a formula."
ws["E3"].font = small_grey

# 6 tiles in 2 rows of 3, using scorecard rows 8-19
# Col groups: B-D, F-H, J-L
tile_map = [
    (2, 8, "Net revenue"),
    (6, 10, "EBITDA"),
    (10, 12, "OTIF %"),
    (2, 15, "CCC (days)"),
    (6, 16, "Headcount"),
    (10, 18, "Op. cash flow"),
]
# row start 5 for first 3, row 11 for next 3
for idx, (coln, sc_row, title) in enumerate(tile_map):
    base = 5 if idx < 3 else 11
    ws.merge_cells(start_row=base, start_column=coln, end_row=base, end_column=coln + 2)
    head = ws.cell(base, coln, title)
    head.font = header_font
    head.fill = header_fill
    head.alignment = Alignment(horizontal="center")
    ws.cell(base, coln + 1).fill = header_fill
    ws.cell(base, coln + 2).fill = header_fill
    ws.cell(base + 1, coln, "Actual").fill = tile_fill
    ws.cell(base + 1, coln).border = thin
    a = ws.cell(base + 1, coln + 1, f"='{SC}'!C{sc_row}")
    style_formula(a, key=True)
    # number formats from KPI
    kpi_idx = sc_row - 8
    a.number_format = KPIS[kpi_idx]["nf"]
    rag = ws.cell(base + 1, coln + 2, f"='{SC}'!G{sc_row}")
    style_formula(rag)
    rag.font = rag_font
    ws.cell(base + 2, coln, "Target").fill = tile_fill
    ws.cell(base + 2, coln).border = thin
    t = ws.cell(base + 2, coln + 1, f"='{SC}'!D{sc_row}")
    style_formula(t)
    t.number_format = KPIS[kpi_idx]["nf"]
    ws.cell(base + 3, coln, "Var / YTD RAG").fill = tile_fill
    ws.cell(base + 3, coln).border = thin
    v = ws.cell(base + 3, coln + 1, f"='{SC}'!E{sc_row}")
    style_formula(v)
    v.number_format = KPIS[kpi_idx]["nf"]
    y = ws.cell(base + 3, coln + 2, f"='{SC}'!K{sc_row}")
    style_formula(y)
    y.font = rag_font

for rng in ["D6:D8", "H6:H8", "L6:L8", "D12:D14", "H12:H14", "L12:L14",
            "D8", "H8", "L8", "D14", "H14", "L14"]:
    pass
for addr in ["D6", "H6", "L6", "D12", "H12", "L12", "D8", "H8", "L8", "D14", "H14", "L14"]:
    ws.conditional_formatting.add(addr, CellIsRule(operator="equal", formula=['"Green"'], fill=green_fill))
    ws.conditional_formatting.add(addr, CellIsRule(operator="equal", formula=['"Amber"'], fill=amber_fill))
    ws.conditional_formatting.add(addr, CellIsRule(operator="equal", formula=['"Red"'], fill=red_fill))

# Mini RAG strip for all 12
ws["B16"] = "Full as-of RAG strip"
ws["B16"].font = section_font
ws["B16"].fill = section_fill
for i, k in enumerate(KPIS):
    c = 3 + i
    ws.cell(17, c, k["name"])
    ws.cell(17, c).font = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
    ws.cell(17, c).fill = header_fill
    ws.cell(17, c).alignment = Alignment(horizontal="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(c)].width = max(12, ws.column_dimensions[get_column_letter(c)].width or 12)
    cell = ws.cell(18, c, f"='{SC}'!G{8+i}")
    style_formula(cell)
    cell.font = rag_font
ws.conditional_formatting.add("C18:N18", CellIsRule(operator="equal", formula=['"Green"'], fill=green_fill))
ws.conditional_formatting.add("C18:N18", CellIsRule(operator="equal", formula=['"Amber"'], fill=amber_fill))
ws.conditional_formatting.add("C18:N18", CellIsRule(operator="equal", formula=['"Red"'], fill=red_fill))
ws.row_dimensions[17].height = 28

# Chart data
ws["B20"] = "Chart data"
ws["B20"].font = section_font
style_header_cell(ws.cell(21, 2), "Metric")
for i, m in enumerate(MONTHS):
    style_header_cell(ws.cell(21, 3 + i), m)
ws["B22"] = "Revenue actual"
ws["B23"] = "Revenue target"
ws["B24"] = "OTIF actual"
ws["B25"] = "OTIF target"
ws["B26"] = "CCC actual"
ws["B27"] = "CCC target"
for m in range(12):
    cl = col(m)
    c = 3 + m
    ws.cell(22, c, f"='{ACT}'!{cl}7"); style_formula(ws.cell(22, c)); ws.cell(22, c).number_format = money
    ws.cell(23, c, f"='{ASSUMP}'!{cl}29"); style_formula(ws.cell(23, c)); ws.cell(23, c).number_format = money
    ws.cell(24, c, f"='{ACT}'!{cl}11"); style_formula(ws.cell(24, c)); ws.cell(24, c).number_format = pct
    ws.cell(25, c, f"='{ASSUMP}'!{cl}33"); style_formula(ws.cell(25, c)); ws.cell(25, c).number_format = pct
    ws.cell(26, c, f"='{ACT}'!{cl}14"); style_formula(ws.cell(26, c)); ws.cell(26, c).number_format = days_fmt
    ws.cell(27, c, f"='{ASSUMP}'!{cl}36"); style_formula(ws.cell(27, c)); ws.cell(27, c).number_format = days_fmt

chart1 = LineChart()
chart1.title = "Revenue actual vs target ($000s)"
chart1.style = 10
chart1.y_axis.title = "$000s"
chart1.height = 7
chart1.width = 15
chart1.legend.position = "b"
data = Reference(ws, min_col=2, min_row=22, max_col=14, max_row=23)
cats = Reference(ws, min_col=3, min_row=21, max_col=14)
chart1.add_data(data, from_rows=True, titles_from_data=True)
chart1.set_categories(cats)
ws.add_chart(chart1, "B29")

chart2 = LineChart()
chart2.title = "OTIF % vs target"
chart2.style = 12
chart2.y_axis.title = "%"
chart2.height = 7
chart2.width = 12
chart2.legend.position = "b"
data2 = Reference(ws, min_col=2, min_row=24, max_col=14, max_row=25)
chart2.add_data(data2, from_rows=True, titles_from_data=True)
chart2.set_categories(cats)
chart2.y_axis.scaling.min = 0.92
ws.add_chart(chart2, "I29")

ws["B45"] = "Park the reporting month on Assumptions to walk the year. Red tiles are the conversation starters; the heatmap on 03_Scorecard shows whether a miss is a one-month blip."
ws["B45"].font = small_grey

# ========== 06_Data_Dictionary ==========
ws = wb.create_sheet("06_Data_Dictionary")
ws.sheet_properties.tabColor = "7F7F7F"
set_col_widths(ws, [4, 32, 22, 78])
landscape(ws)
ws["B2"] = "Data dictionary"
ws["B2"].font = title_font
ws["B3"] = "Field definitions so another analyst can inherit the file."
ws["B3"].font = italic_grey
headers = ["Field", "Tab", "Definition"]
for i, h in enumerate(headers):
    cell = ws.cell(5, 2 + i, h)
    cell.font = header_font; cell.fill = header_fill; cell.border = thin

defs = [
    ("Reporting month", ASSUMP, "1–12 toggle. Scorecard as-of, YTD, and dashboard INDEX to this month."),
    ("Direction H/L/R", ASSUMP, "H = higher-better, L = lower-better, R = range (hit the plan). Drives RAG."),
    ("Amber band", ASSUMP, "Relative miss still scored Amber. H: actual ≥ target×(1−band). L: actual ≤ target×(1+band)."),
    ("Green band (R)", ASSUMP, "For range KPIs only: |actual/target − 1| ≤ green band is Green."),
    ("Aggregation", ASSUMP, "SUM (flows), AVG (rates), LAST (stocks). Used for FY totals and YTD."),
    ("Monthly targets", ASSUMP, "Yellow inputs. Seasonal for revenue/volume/EBITDA; flat policy for OTIF, scrap, DSO, CCC."),
    ("FY25 actuals", ASSUMP, "Prior-year monthly revenue and volume so Trends can show YoY."),
    ("Monthly actuals", ACT, "Yellow inputs — the only numbers a controller would paste each month."),
    ("YTD actual / target", ACT, "OFFSET/INDEX through the reporting month, using each KPI’s aggregation."),
    ("As-of RAG", SC, "Green / Amber / Red from actual vs target and the KPI’s direction + bands. Formula, not typed."),
    ("Monthly RAG heatmap", SC, "Same RAG formula across 12 months. Conditional formatting colors the cells."),
    ("MoM change", TR, "For rates, actual minus prior month. For levels, actual / prior − 1."),
    ("Rolling 3-month", TR, "Average of the last 1–3 months (shorter window at the start of the year)."),
    ("YoY %", TR, "FY26 actual ÷ FY25 actual − 1 for revenue and volume."),
    ("Net revenue", ACT, "Net sales $000s. Higher-better."),
    ("Gross margin %", ACT, "(Revenue − COGS) / revenue. Higher-better."),
    ("EBITDA", ACT, "Earnings before interest, tax, depreciation, amortization. $000s. Higher-better."),
    ("Volume", ACT, "Cases shipped, thousands. Higher-better."),
    ("OTIF %", ACT, "On-time in-full customer deliveries. Higher-better."),
    ("Scrap %", ACT, "Scrapped production ÷ total production. Lower-better."),
    ("DSO / CCC", ACT, "Working-capital days. Lower-better. Aligns with the WC dashboard sample."),
    ("Headcount", ACT, "Ending FTE. Range vs plan (R). Aligns with the staffing-forecast sample."),
    ("Attrition % (ann.)", ACT, "Monthly exits / opening × 12. Lower-better."),
    ("Operating cash flow", ACT, "Cash from operations $000s. Higher-better."),
    ("SG&A % of sales", ACT, "Selling, general & administrative ÷ net revenue. Lower-better."),
    ("Units", "All", "Dollars in $000s. Rates in %. Headcount in FTE. Volume in 000 cases. Days in days."),
]
for i, (field, tab, definition) in enumerate(defs):
    r = 6 + i
    ws.cell(r, 2, field).font = bold
    ws.cell(r, 2).border = thin
    ws.cell(r, 3, tab).border = thin
    cell = ws.cell(r, 4, definition)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.border = thin
    ws.row_dimensions[r].height = 32

ws["B34"] = "All sample numbers are fictional. Built for a public GitHub portfolio — no employer data."
ws["B34"].font = small_grey

out = Path(__file__).resolve().parent / "Northline_Monthly_KPI_Scorecard.xlsx"
wb.save(out)
print("Wrote", out)
print("Sheets:", wb.sheetnames)
